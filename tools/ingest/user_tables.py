"""Attach operator-supplied tabular metadata to the sidecar (`user_provided_metadata`).

Some batches arrive with an accompanying spreadsheet: a collaborator's own
table that carries per-acquisition facts the instrument files do not
(contributing centre, per-case caveats, the grant the data was originally
collected under). This module reads such a table and nests it in the
`metadata.json` sidecar under `user_provided_metadata` — see
[08_METADATA §4.8](../../mfb-rdm-docs/08_METADATA.md).

Why its own block, and why namespaced by label:

  * These values are **asserted by a person**, not computed by the pipeline.
    Several columns (file size, checksum-present, canonical path) deliberately
    shadow registry fields the ingest derives itself, and the two can disagree.
    Keeping them under `user_provided_metadata.<label>` means a reader always
    knows which is which — the registry stays the single source of computed
    truth, and the collaborator's claim is preserved verbatim as provenance.
  * The label is chosen in the YAML, not taken from the filename, so two
    cohorts of one project that ship the *same* logical table under different
    filenames (`dataset_information_HPIC.xlsx` / `..._lions.xlsx`) land on the
    same key and stay queryable as one field. The exact filename, sheet and
    matched row are preserved in `_source`.

SCOPE (deliberate, 2026-08-12). This is a flat, per-acquisition attachment.
It is NOT the right home for two things it will be tempting to use it for:
study-level structure (the ISA investigation/study/assay hierarchy — see
BACKLOG META-10) and new *measurement* data such as clinical hemodynamics,
which is really its own acquisition of its own data type linked by subject
(BACKLOG META-11). Both are recorded as backlog items rather than solved here.

Two table orientations are supported:

  row      (default) One row per acquisition. The case is located by matching
           a resolved `match:` expression against `key_column`.
  vertical A Field/Value table that describes the WHOLE batch (e.g. the
           project the data was originally collected under). Every
           acquisition in the batch gets the same block.

Workbooks are parsed once and memoized per (path, sheet, mtime), so a 42-case
batch reads each spreadsheet a single time.
"""

import csv
import os
from datetime import date, datetime

from . import resolver


class UserTableError(Exception):
    """Raised for a malformed `user_metadata:` block or an unreadable table."""


# Cache: (abspath, sheet, mtime, orientation-ish key) -> parsed table
_TABLE_CACHE = {}

VALID_ORIENTATIONS = ("row", "vertical")
VALID_ON_MISSING = ("warn", "error", "skip")

# Keys accepted inside one `user_metadata:` list entry.
_VALID_KEYS = {
    "label", "file", "sheet", "orientation", "header_row",
    "key_column", "match", "key_transform", "match_transform",
    "field_column", "value_column", "description_column", "skip_columns",
    "split_descriptions", "on_missing",
}


# --------------------------------------------------------------------------
# value / key normalization
# --------------------------------------------------------------------------

def _clean_text(value):
    """Normalize a header or key cell: NBSP -> space, collapse, strip."""
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def _json_value(value):
    """Coerce an Excel cell to something json.dump can write.

    Dates become ISO strings; numbers and bools pass through; everything else
    is stringified and whitespace-normalized. Empty cells become None so a
    blank column is visibly blank rather than silently absent.
    """
    if value is None:
        return None
    if isinstance(value, bool) or isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        # Midnight-only timestamps are dates in every sheet we have seen.
        if value.hour or value.minute or value.second:
            return value.strftime("%Y-%m-%dT%H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = _clean_text(value)
    return text or None


def apply_transform(value, transform):
    """Apply one join-key transform. Returns a string.

    Transforms exist because the join key rarely matches on both sides
    verbatim:

      decimal2      Excel stores an ID like `1.10` as the NUMBER 1.1, which
                    stringifies to "1.1" and silently fails to match the
                    `LEONE_1.10` case folder. Re-pads to two decimals. Only
                    touches real numbers, so text IDs pass through untouched.
      first_token   "HPIC37 S63090" -> "HPIC37" (folder carries an extra
                    accession suffix the table does not).
      strip_prefix:<p>  "LEONE_1.01" -> "1.01".
      lower / upper / strip
    """
    if transform is None or transform == "":
        return value if isinstance(value, str) else _clean_text(value)

    if transform == "decimal2":
        if isinstance(value, bool):
            return _clean_text(value)
        if isinstance(value, (int, float)):
            return f"{value:.2f}"
        return _clean_text(value)

    text = value if isinstance(value, str) else _clean_text(value)
    if transform == "first_token":
        parts = text.split()
        return parts[0] if parts else ""
    if transform.startswith("strip_prefix:"):
        prefix = transform.split(":", 1)[1]
        return text[len(prefix):] if prefix and text.startswith(prefix) else text
    if transform == "lower":
        return text.lower()
    if transform == "upper":
        return text.upper()
    if transform == "strip":
        return text.strip()
    raise UserTableError(
        f"user_metadata: unknown transform {transform!r}. Valid: decimal2, "
        f"first_token, strip_prefix:<p>, lower, upper, strip"
    )


# --------------------------------------------------------------------------
# validation
# --------------------------------------------------------------------------

def validate_user_metadata_block(block):
    """Return a list of error strings for a `user_metadata:` block ([] if OK).

    Validated at config-load time so a typo fails before any data is copied.
    """
    errors = []
    if block is None:
        return errors
    if not isinstance(block, list):
        return ["user_metadata: must be a list of table entries"]

    seen_labels = set()
    for i, entry in enumerate(block):
        where = f"user_metadata[{i}]"
        if not isinstance(entry, dict):
            errors.append(f"{where}: must be a mapping")
            continue

        unknown = sorted(set(entry) - _VALID_KEYS)
        if unknown:
            errors.append(
                f"{where}: unknown key(s) {', '.join(unknown)}. "
                f"Valid: {', '.join(sorted(_VALID_KEYS))}"
            )

        label = entry.get("label")
        if not label:
            errors.append(f"{where}: 'label' is required (the sidecar key)")
        elif label in seen_labels:
            errors.append(f"{where}: duplicate label {label!r}")
        else:
            seen_labels.add(label)
        if label and str(label).startswith("_"):
            errors.append(f"{where}: label may not start with '_' (reserved)")

        path = entry.get("file")
        if not path:
            errors.append(f"{where}: 'file' is required")
        elif not os.path.isfile(path):
            errors.append(f"{where}: file not found: {path}")

        orientation = (entry.get("orientation") or "row").lower()
        if orientation not in VALID_ORIENTATIONS:
            errors.append(
                f"{where}: orientation must be one of {VALID_ORIENTATIONS}"
            )
        elif orientation == "row":
            if not entry.get("key_column"):
                errors.append(f"{where}: 'key_column' is required for orientation: row")
            if not entry.get("match"):
                errors.append(f"{where}: 'match' is required for orientation: row")

        on_missing = (entry.get("on_missing") or "warn").lower()
        if on_missing not in VALID_ON_MISSING:
            errors.append(f"{where}: on_missing must be one of {VALID_ON_MISSING}")

        header_row = entry.get("header_row", 1)
        if not isinstance(header_row, int) or header_row < 1:
            errors.append(f"{where}: header_row must be a positive integer (1-based)")

        skip = entry.get("skip_columns")
        if skip is not None and not isinstance(skip, list):
            errors.append(f"{where}: skip_columns must be a list of column names")

        for key in ("key_transform", "match_transform"):
            t = entry.get(key)
            if t:
                try:
                    apply_transform("probe", t)
                except UserTableError as exc:
                    errors.append(f"{where}: {exc}")
    return errors


# --------------------------------------------------------------------------
# reading
# --------------------------------------------------------------------------

def _read_grid(path, sheet):
    """Return the sheet (or CSV) as a list of row-lists of raw cell values."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as fh:
            return [list(r) for r in csv.reader(fh)]
    if ext not in (".xlsx", ".xlsm"):
        raise UserTableError(
            f"user_metadata: unsupported table format {ext!r} for {path} "
            f"(supported: .xlsx, .xlsm, .csv)"
        )
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment issue
        raise UserTableError(
            "user_metadata: reading .xlsx requires openpyxl (pip install openpyxl)"
        ) from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet:
            if sheet not in wb.sheetnames:
                raise UserTableError(
                    f"user_metadata: sheet {sheet!r} not in {os.path.basename(path)} "
                    f"(has: {', '.join(wb.sheetnames)})"
                )
            ws = wb[sheet]
        else:
            ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _parse_table(entry):
    """Parse one `user_metadata:` entry into a lookup structure.

    Returns a dict:
        orientation   "row" | "vertical"
        fields        ordered list of field names
        descriptions  {field: description} when split_descriptions is on
        rows          {normalized_key: {field: value}}     (row orientation)
        constant      {field: value}                       (vertical)
        source        {file, sheet, ...} provenance stub
    """
    path = entry["file"]
    sheet = entry.get("sheet")
    orientation = (entry.get("orientation") or "row").lower()
    cache_key = (os.path.abspath(path), sheet, os.path.getmtime(path),
                 orientation, entry.get("header_row", 1),
                 entry.get("key_column"), entry.get("key_transform"),
                 tuple(entry.get("skip_columns") or ()),
                 bool(entry.get("split_descriptions")),
                 entry.get("field_column"), entry.get("value_column"))
    if cache_key in _TABLE_CACHE:
        return _TABLE_CACHE[cache_key]

    grid = _read_grid(path, sheet)
    source = {"file": os.path.basename(path)}
    if sheet:
        source["sheet"] = sheet

    if orientation == "vertical":
        table = _parse_vertical(entry, grid, source)
    else:
        table = _parse_row(entry, grid, source)
    _TABLE_CACHE[cache_key] = table
    return table


def _split_header(raw, split_descriptions):
    """Split a header cell into (field_name, description).

    The dataset_information sheets embed the field's own documentation in the
    header cell, e.g. `operator " Person who collected the data ..."`. With
    split_descriptions on, the name is everything before the first quote and
    the remainder is kept as the field's description.
    """
    text = _clean_text(raw)
    if not split_descriptions or not text:
        return text, None
    for quote in ('"', '“', '”'):
        if quote in text:
            name, _, rest = text.partition(quote)
            name = name.strip()
            desc = rest.strip().strip('"“”').strip()
            if name:
                return name, (desc or None)
    return text, None


def _parse_row(entry, grid, source):
    header_idx = entry.get("header_row", 1) - 1
    if header_idx >= len(grid):
        raise UserTableError(
            f"user_metadata[{entry.get('label')}]: header_row "
            f"{entry.get('header_row', 1)} is past the end of the sheet "
            f"({len(grid)} rows)"
        )
    split_desc = bool(entry.get("split_descriptions"))
    skip = {_clean_text(c) for c in (entry.get("skip_columns") or [])}

    fields, descriptions = [], {}
    for cell in grid[header_idx]:
        name, desc = _split_header(cell, split_desc)
        fields.append(name)
        if desc:
            descriptions[name] = desc

    key_column = _clean_text(entry["key_column"])
    if key_column not in fields:
        raise UserTableError(
            f"user_metadata[{entry.get('label')}]: key_column {key_column!r} "
            f"not among the header fields {[f for f in fields if f]}"
        )
    key_idx = fields.index(key_column)
    key_transform = entry.get("key_transform")

    rows, raw_keys, dup_keys, float_key_seen = {}, {}, set(), False
    for raw_row in grid[header_idx + 1:]:
        if key_idx >= len(raw_row):
            continue
        raw_key = raw_row[key_idx]
        if raw_key is None or _clean_text(raw_key) == "":
            continue
        if isinstance(raw_key, float) and not key_transform:
            float_key_seen = True
        key = apply_transform(raw_key, key_transform)
        if key in rows:
            dup_keys.add(key)
            continue
        record = {}
        for col_idx, field in enumerate(fields):
            if not field or field in skip:
                continue
            value = raw_row[col_idx] if col_idx < len(raw_row) else None
            record[field] = _json_value(value)
        # The key column carries the NORMALIZED key, not Excel's raw cell:
        # a sheet that stored `1.10` as the number 1.1 would otherwise write
        # `"AcquisitionID": 1.1` next to a `matched_on.value` of "1.10" and
        # read like a mismatch. The raw cell is preserved in _source below,
        # so nothing is lost.
        raw_key_value = _json_value(raw_key)
        if key_transform and record.get(key_column) != key:
            record[key_column] = key
        rows[key] = record
        raw_keys[key] = raw_key_value

    return {
        "orientation": "row",
        "fields": fields,
        "descriptions": descriptions,
        "rows": rows,
        "raw_keys": raw_keys,
        "source": source,
        "key_column": key_column,
        "duplicate_keys": sorted(dup_keys),
        "float_key_seen": float_key_seen,
    }


def _parse_vertical(entry, grid, source):
    """Parse a Field/Value table that applies to the whole batch."""
    header_idx = entry.get("header_row", 1) - 1
    field_col = entry.get("field_column", 1)
    value_col = entry.get("value_column", 2)
    desc_col = entry.get("description_column")
    split_desc = bool(entry.get("split_descriptions"))
    skip = {_clean_text(c) for c in (entry.get("skip_columns") or [])}

    def _col_index(spec, header):
        """A column is addressed by 1-based position or by header text."""
        if isinstance(spec, int):
            return spec - 1
        wanted = _clean_text(spec)
        cleaned = [_clean_text(c) for c in header]
        if wanted in cleaned:
            return cleaned.index(wanted)
        raise UserTableError(
            f"user_metadata[{entry.get('label')}]: column {spec!r} not found "
            f"in header {[c for c in cleaned if c]}"
        )

    header = grid[header_idx] if header_idx < len(grid) else []
    f_idx = _col_index(field_col, header)
    v_idx = _col_index(value_col, header)
    d_idx = _col_index(desc_col, header) if desc_col is not None else None

    constant, descriptions = {}, {}
    for raw_row in grid[header_idx + 1:]:
        if f_idx >= len(raw_row):
            continue
        name, desc = _split_header(raw_row[f_idx], split_desc)
        if not name or name in skip:
            continue
        value = raw_row[v_idx] if v_idx < len(raw_row) else None
        constant[name] = _json_value(value)
        if d_idx is not None and d_idx < len(raw_row):
            desc = _clean_text(raw_row[d_idx]) or desc
        if desc:
            descriptions[name] = desc
    return {
        "orientation": "vertical",
        "fields": list(constant),
        "descriptions": descriptions,
        "constant": constant,
        "source": source,
    }


# --------------------------------------------------------------------------
# per-acquisition assembly
# --------------------------------------------------------------------------

def build_user_metadata(block, discovered, log=None, dry_run=False):
    """Build the sidecar's `user_provided_metadata` dict for one acquisition.

    Args:
        block: the config's `user_metadata:` list (None/[] -> returns None).
        discovered: the case's discovered dict, for resolving `match:`.
        log: optional log(msg, level) callable.
        dry_run: unused today; accepted so the call site reads like its peers.

    Returns:
        dict for the sidecar, or None when the block is absent or nothing
        matched (so sidecars for batches without a table are unchanged).

    Non-blocking by default (08_METADATA §4.7): a case with no matching row
    WARNs and is omitted from the block rather than failing the ingest. Set
    `on_missing: error` on a table that must match every case.
    """
    def _log(msg, level="INFO"):
        if log:
            log(msg, level)

    if not block:
        return None

    out = {}
    for entry in block:
        label = entry["label"]
        orientation = (entry.get("orientation") or "row").lower()
        on_missing = (entry.get("on_missing") or "warn").lower()
        table = _parse_table(entry)

        if orientation == "vertical":
            record = dict(table["constant"])
            source = dict(table["source"])
            source["orientation"] = "vertical"
        else:
            match_raw = resolver.resolve_value(
                entry["match"], discovered, key_for_error=f"user_metadata.{label}.match"
            )
            match_key = apply_transform(match_raw, entry.get("match_transform"))
            record = table["rows"].get(match_key)
            if record is None:
                msg = (
                    f"user_metadata[{label}]: no row in "
                    f"{table['source']['file']} where "
                    f"{table['key_column']} == {match_key!r}"
                )
                if on_missing == "error":
                    raise UserTableError(msg)
                if on_missing == "warn":
                    hint = ""
                    if table.get("float_key_seen"):
                        hint = (" (the sheet stores some keys as numbers — "
                                "try key_transform: decimal2)")
                    _log(msg + hint, "WARN")
                continue
            record = dict(record)
            source = dict(table["source"])
            source["matched_on"] = {
                "column": table["key_column"],
                "value": match_key,
            }
            raw_key = table["raw_keys"].get(match_key)
            if raw_key is not None and str(raw_key) != match_key:
                source["matched_on"]["raw_value"] = raw_key
            if table["duplicate_keys"]:
                source["duplicate_keys_ignored"] = table["duplicate_keys"]

        if table["descriptions"]:
            source["field_descriptions"] = table["descriptions"]

        block_out = {"_source": source}
        block_out.update(record)
        out[label] = block_out

    return out or None
